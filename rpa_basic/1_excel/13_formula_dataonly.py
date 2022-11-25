from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

# 수식 그대로 가져오고 있음
for row in ws.values:
    for cell in row:
        print(cell)

print()

wb2 = load_workbook("sample_formula.xlsx", data_only=True)
ws2 = wb2.active
# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None 으로 표시/ 파일 열었다가 다시 닫으면 계산된 결과 받아옴
for row in ws2.values:
    for cell in row:
        print(cell)


